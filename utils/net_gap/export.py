# utils/net_gap/export.py - VERSION 4.5

"""
Excel Export - VERSION 4.5
Synchronized with new status classification:
- net_gap < 0 → SHORTAGE (always!)
- net_gap = 0 → BALANCED
- net_gap > 0 → SURPLUS (always!)
"""

import pandas as pd
import numpy as np
from typing import Dict, Any, Optional
from datetime import datetime
import io
import logging

from .constants import EXPORT_CONFIG, GAP_CATEGORIES, THRESHOLDS
from .formatters import GAPFormatter

logger = logging.getLogger(__name__)


def export_to_excel(
    result,
    filters: Dict[str, Any],
    include_cost_breakdown: bool = True
) -> bytes:
    """Export GAP analysis to Excel - v4.5"""
    
    formatter = GAPFormatter()
    output = io.BytesIO()
    
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # 1. Summary Sheet
            summary_df = _create_summary_sheet(result, filters, formatter)
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # 2. GAP Details
            include_safety = filters.get('include_safety', False)
            include_expired = filters.get('include_expired', False)
            details_df = _create_details_sheet(
                result.gap_df, 
                formatter,
                include_safety=include_safety,
                include_expired=include_expired
            )
            details_df.to_excel(writer, sheet_name='GAP Details', index=False)
            
            # 3. Cost Breakdown
            if include_cost_breakdown and 'avg_unit_cost_usd' in result.gap_df.columns:
                cost_df = _create_cost_breakdown(result.gap_df, formatter)
                cost_df.to_excel(writer, sheet_name='Cost Analysis', index=False)
            
            # 4. Calculation Guide - v4.5 logic
            guide_df = _create_calculation_guide(include_safety)
            guide_df.to_excel(writer, sheet_name='Calculation Guide', index=False)
            
            # 5. Customer Impact
            if result.customer_impact and not result.customer_impact.is_empty():
                customer_df = _create_customer_sheet(result.customer_impact)
                customer_df.to_excel(writer, sheet_name='Customer Impact', index=False)
            
            _format_excel_sheets(writer)
        
        output.seek(0)
        logger.info("Excel export generated successfully (v4.5)")
        return output.getvalue()
        
    except Exception as e:
        logger.error(f"Excel export failed: {e}", exc_info=True)
        raise


def _create_summary_sheet(result, filters: Dict, formatter) -> pd.DataFrame:
    """Create summary sheet"""
    
    metrics = result.metrics
    categories = result.get_category_summary()
    
    summary_data = []
    
    # Report Info
    summary_data.extend([
        ['Report Information', ''],
        ['Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ['Version', 'v4.5 - Net GAP Sign-based Classification'],
        ['', '']
    ])
    
    # Filters Applied
    summary_data.extend([
        ['Filters Applied', ''],
        ['Entity', filters.get('entity', 'All Entities')],
        ['Products', f"{len(filters.get('products', []))} selected" if filters.get('products') else 'All'],
        ['Brands', ', '.join(filters.get('brands', [])) or 'All'],
        ['Expired Inventory', 'Excluded' if filters.get('exclude_expired') else 'Included'],
        ['Safety Stock', 'Included' if filters.get('include_safety') else 'Not Included'],
        ['', '']
    ])
    
    # Key Metrics
    summary_data.extend([
        ['Key Metrics', ''],
        ['Total Products', f"{metrics['total_products']:,}"],
        ['Coverage Rate', f"{metrics['overall_coverage']:.1f}%"],
        ['', ''],
        ['Shortage Items', f"{metrics['shortage_items']:,}"],
        ['Optimal Items', f"{categories.get('optimal', 0):,}"],
        ['Surplus Items', f"{metrics['surplus_items']:,}"],
        ['Inactive Items', f"{categories.get('inactive', 0):,}"],
        ['', '']
    ])
    
    # Financial Impact
    summary_data.extend([
        ['Financial Impact', ''],
        ['Total Supply Value', formatter.format_currency(metrics.get('total_supply_value_usd', 0))],
        ['Total Demand Value', formatter.format_currency(metrics.get('total_demand_value_usd', 0))],
        ['Revenue at Risk', formatter.format_currency(metrics['at_risk_value_usd'])],
        ['Total Shortage Qty', f"{metrics['total_shortage']:,.0f} units"],
        ['Total Surplus Qty', f"{metrics['total_surplus']:,.0f} units"],
        ['', '']
    ])
    
    # Customer Impact
    if metrics.get('affected_customers', 0) > 0:
        summary_data.extend([
            ['Customer Impact', ''],
            ['Affected Customers', f"{metrics['affected_customers']:,}"]
        ])
    
    # Safety Stock Metrics
    if filters.get('include_safety') and 'below_safety_count' in metrics:
        summary_data.extend([
            ['', ''],
            ['Safety Stock Metrics', ''],
            ['Below Safety Count', f"{metrics.get('below_safety_count', 0):,}"],
            ['At Reorder Count', f"{metrics.get('at_reorder_count', 0):,}"]
        ])
    
    return pd.DataFrame(summary_data, columns=['Metric', 'Value'])


def _create_details_sheet(
    gap_df: pd.DataFrame, 
    formatter,
    include_safety: bool = False,
    include_expired: bool = False
) -> pd.DataFrame:
    """Create details sheet synchronized with UI"""
    
    if gap_df.empty:
        return pd.DataFrame()
    
    export_df = pd.DataFrame()
    
    # Product identification
    if 'pt_code' in gap_df.columns:
        export_df['PT Code'] = gap_df['pt_code']
    if 'product_name' in gap_df.columns:
        export_df['Product Name'] = gap_df['product_name']
    if 'brand' in gap_df.columns:
        export_df['Brand'] = gap_df['brand']
    if 'standard_uom' in gap_df.columns:
        export_df['UOM'] = gap_df['standard_uom']
    
    # Supply columns
    if 'total_supply' in gap_df.columns:
        export_df['Total Supply'] = gap_df['total_supply'].round(2)
    
    if 'supply_inventory' in gap_df.columns:
        export_df['Inventory'] = gap_df['supply_inventory'].round(2)
    if 'supply_can_pending' in gap_df.columns:
        export_df['CAN Pending'] = gap_df['supply_can_pending'].round(2)
    if 'supply_warehouse_transfer' in gap_df.columns:
        export_df['Warehouse Transfer'] = gap_df['supply_warehouse_transfer'].round(2)
    if 'supply_purchase_order' in gap_df.columns:
        export_df['Purchase Order'] = gap_df['supply_purchase_order'].round(2)
    
    # Demand columns
    if 'total_demand' in gap_df.columns:
        export_df['Total Demand'] = gap_df['total_demand'].round(2)
    
    if 'demand_oc_pending' in gap_df.columns:
        export_df['OC Pending'] = gap_df['demand_oc_pending'].round(2)
    if 'demand_forecast' in gap_df.columns:
        export_df['Forecast'] = gap_df['demand_forecast'].round(2)
    
    # Safety Stock columns
    if include_safety:
        if 'safety_stock_qty' in gap_df.columns:
            export_df['Safety Stock'] = gap_df['safety_stock_qty'].round(2)
        
        if 'safety_gap' in gap_df.columns:
            export_df['Safety Gap'] = gap_df['safety_gap'].round(2)
        
        if 'available_supply' in gap_df.columns:
            export_df['Available Supply'] = gap_df['available_supply'].round(2)
    
    # GAP columns
    if 'net_gap' in gap_df.columns:
        export_df['Net GAP'] = gap_df['net_gap'].round(2)
    
    if 'true_gap' in gap_df.columns:
        export_df['True GAP'] = gap_df['true_gap'].round(2)
    
    # Shortage Cause
    if 'shortage_cause' in gap_df.columns:
        export_df['Shortage Cause'] = gap_df['shortage_cause'].apply(_remove_emojis)
    
    # Coverage
    if 'coverage_ratio' in gap_df.columns:
        export_df['Coverage %'] = gap_df['coverage_ratio'].apply(
            lambda x: f"{x*100:.1f}%" if pd.notna(x) else "N/A"
        )
    
    if 'gap_percentage' in gap_df.columns:
        export_df['GAP %'] = gap_df['gap_percentage'].apply(
            lambda x: f"{x:+.1f}%" if pd.notna(x) else "N/A"
        )
    
    # Additional Safety columns
    if include_safety:
        if 'reorder_point' in gap_df.columns:
            export_df['Reorder Point'] = gap_df['reorder_point'].round(2)
        
        if 'below_reorder' in gap_df.columns:
            export_df['Below Reorder'] = gap_df['below_reorder'].apply(
                lambda x: 'Yes' if x else 'No'
            )
        
        if 'safety_coverage' in gap_df.columns:
            export_df['Safety Coverage'] = gap_df['safety_coverage'].apply(
                lambda x: f"{x:.1f}x" if pd.notna(x) and x < 999 else "N/A"
            )
    
    # Financial columns
    if 'avg_unit_cost_usd' in gap_df.columns:
        export_df['Unit Cost (USD)'] = gap_df['avg_unit_cost_usd'].round(4)
    
    if 'avg_selling_price_usd' in gap_df.columns:
        export_df['Sell Price (USD)'] = gap_df['avg_selling_price_usd'].round(4)
    
    if 'supply_value_usd' in gap_df.columns:
        export_df['Supply Value (USD)'] = gap_df['supply_value_usd'].round(2)
    
    if 'demand_value_usd' in gap_df.columns:
        export_df['Demand Value (USD)'] = gap_df['demand_value_usd'].round(2)
    
    if 'gap_value_usd' in gap_df.columns:
        export_df['GAP Value (USD)'] = gap_df['gap_value_usd'].round(2)
    
    if 'at_risk_value_usd' in gap_df.columns:
        export_df['At Risk Value (USD)'] = gap_df['at_risk_value_usd'].round(2)
    
    # Status columns
    if 'gap_status' in gap_df.columns:
        export_df['Status'] = gap_df['gap_status'].apply(
            lambda x: x.replace('_', ' ').title() if pd.notna(x) else ''
        )
    
    if 'priority' in gap_df.columns:
        priority_map = {1: 'P1-Critical', 2: 'P2-High', 3: 'P3-Medium', 4: 'P4-Low', 99: 'P99-OK'}
        export_df['Priority'] = gap_df['priority'].map(priority_map).fillna('Unknown')
    
    if 'suggested_action' in gap_df.columns:
        export_df['Suggested Action'] = gap_df['suggested_action'].apply(_remove_emojis)
    
    # Customer count
    if 'customer_count' in gap_df.columns:
        export_df['Customer Count'] = gap_df['customer_count'].fillna(0).astype(int)
    
    # Expired inventory
    if include_expired:
        if 'expired_quantity' in gap_df.columns:
            export_df['Expired Qty'] = gap_df['expired_quantity'].round(2)
        
        if 'expired_batches_info' in gap_df.columns:
            export_df['Expired Batches'] = gap_df['expired_batches_info'].fillna('')
    
    # Product ID
    if 'product_id' in gap_df.columns:
        export_df['Product ID'] = gap_df['product_id']
    
    # Limit rows
    if len(export_df) > EXPORT_CONFIG['max_rows']:
        export_df = export_df.head(EXPORT_CONFIG['max_rows'])
        logger.warning(f"Export limited to {EXPORT_CONFIG['max_rows']} rows")
    
    return export_df


def _remove_emojis(text: str) -> str:
    """Remove emojis from text"""
    if not text:
        return ''
    
    emojis = ['🚨', '⚠️', '📋', '🔒', '✅', '📦', '🛑', '⭕', '⚪', '🔴', '🟠', '🟡', '🔵', '🟣', '🟢', '❓']
    result = text
    for emoji in emojis:
        result = result.replace(emoji, '')
    return result.strip()


def _create_cost_breakdown(gap_df: pd.DataFrame, formatter) -> pd.DataFrame:
    """Create cost breakdown sheet"""
    
    breakdown_data = []
    
    for _, row in gap_df.iterrows():
        unit_cost = row.get('avg_unit_cost_usd', 0) or 0
        sell_price = row.get('avg_selling_price_usd', 0) or 0
        net_gap = row.get('net_gap', 0) or 0
        total_supply = row.get('total_supply', 0) or 0
        total_demand = row.get('total_demand', 0) or 0
        
        breakdown_data.append({
            'Product Code': row.get('pt_code', ''),
            'Product Name': row.get('product_name', ''),
            'Brand': row.get('brand', ''),
            
            'Supply Qty': round(total_supply, 2),
            'Demand Qty': round(total_demand, 2),
            'Net GAP Qty': round(net_gap, 2),
            'True GAP Qty': round(row.get('true_gap', total_supply - total_demand), 2),
            
            'Avg Unit Cost (USD)': round(unit_cost, 4),
            'Avg Selling Price (USD)': round(sell_price, 4),
            'Margin per Unit (USD)': round(sell_price - unit_cost, 4),
            
            'Supply Value (USD)': round(total_supply * unit_cost, 2),
            'Demand Value (USD)': round(total_demand * sell_price, 2),
            'GAP Value (USD)': round(row.get('gap_value_usd', 0) or 0, 2),
            
            'Shortage Qty': round(abs(net_gap), 2) if net_gap < 0 else 0,
            'Revenue at Risk (USD)': round(row.get('at_risk_value_usd', 0) or 0, 2),
            'Lost Margin (USD)': round(abs(net_gap) * (sell_price - unit_cost), 2) if net_gap < 0 else 0,
            
            'Status': row.get('gap_status', ''),
            'Shortage Cause': _remove_emojis(row.get('shortage_cause', '')),
            'Priority': row.get('priority', 99)
        })
    
    return pd.DataFrame(breakdown_data)


def _create_calculation_guide(include_safety: bool = False) -> pd.DataFrame:
    """Create calculation guide - Complete version"""
    
    guide_data = []
    
    # ===== SECTION 1: CORE FORMULAS =====
    guide_data.append({
        'Section': 'CORE FORMULAS',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Safety Gap',
        'Formula': 'Total Supply - Safety Stock',
        'Example': '100 - 20 = 80'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '  - Can be negative when supply below safety',
        'Formula': '',
        'Example': '3 - 25 = -22'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Available Supply',
        'Formula': 'MAX(0, Safety Gap)',
        'Example': 'MAX(0, 80) = 80'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '  - Capped at 0 (never negative)',
        'Formula': '',
        'Example': 'MAX(0, -22) = 0'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Net GAP',
        'Formula': 'Available Supply - Total Demand',
        'Example': '80 - 50 = +30'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '  - Primary metric for shortage/surplus',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'True GAP',
        'Formula': 'Total Supply - Total Demand',
        'Example': '100 - 50 = +50'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '  - Always ignores safety stock',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage Ratio',
        'Formula': '(Available Supply / Demand) x 100%',
        'Example': '(80 / 50) x 100% = 160%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    # ===== SECTION 2: STATUS LOGIC =====
    guide_data.append({
        'Section': 'STATUS LOGIC',
        'Description': 'Net GAP sign determines group',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Net GAP < 0',
        'Formula': '-> SHORTAGE (always!)',
        'Example': '-20 -> Shortage'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Net GAP = 0',
        'Formula': '-> BALANCED',
        'Example': '0 -> Balanced'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Net GAP > 0',
        'Formula': '-> SURPLUS (always!)',
        'Example': '+10 -> Surplus'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    # ===== SECTION 3: SHORTAGE SEVERITY =====
    guide_data.append({
        'Section': 'SHORTAGE SEVERITY',
        'Description': 'Coverage determines severity level',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage < 25%',
        'Formula': '-> CRITICAL SHORTAGE',
        'Example': '20/100 = 20%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage < 50%',
        'Formula': '-> SEVERE SHORTAGE',
        'Example': '45/100 = 45%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage < 75%',
        'Formula': '-> HIGH SHORTAGE',
        'Example': '70/100 = 70%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage < 90%',
        'Formula': '-> MODERATE SHORTAGE',
        'Example': '85/100 = 85%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage < 100%',
        'Formula': '-> LIGHT SHORTAGE',
        'Example': '92/100 = 92%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    # ===== SECTION 4: SURPLUS SEVERITY =====
    guide_data.append({
        'Section': 'SURPLUS SEVERITY',
        'Description': 'Coverage determines severity level',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage <= 125%',
        'Formula': '-> LIGHT SURPLUS',
        'Example': '120/100 = 120%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage <= 175%',
        'Formula': '-> MODERATE SURPLUS',
        'Example': '160/100 = 160%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage <= 250%',
        'Formula': '-> HIGH SURPLUS',
        'Example': '230/100 = 230%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Coverage > 250%',
        'Formula': '-> SEVERE SURPLUS',
        'Example': '300/100 = 300%'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    # ===== SECTION 5: FINANCIAL CALCULATIONS =====
    guide_data.append({
        'Section': 'FINANCIAL CALCULATIONS',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'At Risk Value',
        'Formula': '|Shortage Qty| x Selling Price',
        'Example': '20 x $50 = $1,000'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '  - Revenue at risk from shortage',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'GAP Value',
        'Formula': 'Net GAP x Unit Cost',
        'Example': '-20 x $30 = -$600'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '  - Inventory value of the gap',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Safety Stock Impact',
        'Formula': 'Net GAP - True GAP',
        'Example': '-10 - (+10) = -20'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '  - Negative = safety creates shortage',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    # ===== SECTION 6: SHORTAGE CAUSES =====
    guide_data.append({
        'Section': 'SHORTAGE CAUSES',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'OK - No shortage',
        'Formula': 'Net GAP >= 0',
        'Example': 'Supply meets demand'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Safety Requirement',
        'Formula': 'True GAP >= 0 but Net GAP < 0',
        'Example': 'Shortage from safety reservation'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Real Shortage',
        'Formula': 'True GAP < 0',
        'Example': 'Actual supply < demand'
    })
    
    guide_data.append({
        'Section': '',
        'Description': 'Supply < Safety Req.',
        'Formula': 'Supply below safety stock level',
        'Example': 'Supply=3, Safety=25'
    })
    
    guide_data.append({
        'Section': '',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    # ===== SECTION 7: EXAMPLE SCENARIOS =====
    guide_data.append({
        'Section': 'EXAMPLE SCENARIOS',
        'Description': '',
        'Formula': '',
        'Example': ''
    })
    
    guide_data.append({
        'Section': 'Healthy',
        'Description': 'Supply=100, Safety=20, Demand=50',
        'Formula': 'Safety Gap=+80, Available=80',
        'Example': 'Net GAP=+30, True GAP=+50, Cause=OK'
    })
    
    guide_data.append({
        'Section': 'Tight',
        'Description': 'Supply=100, Safety=20, Demand=90',
        'Formula': 'Safety Gap=+80, Available=80',
        'Example': 'Net GAP=-10, True GAP=+10, Cause=Safety Req.'
    })
    
    guide_data.append({
        'Section': 'Real Shortage',
        'Description': 'Supply=50, Safety=20, Demand=80',
        'Formula': 'Safety Gap=+30, Available=30',
        'Example': 'Net GAP=-50, True GAP=-30, Cause=Real Shortage'
    })
    
    guide_data.append({
        'Section': 'Under Safety',
        'Description': 'Supply=3, Safety=25, Demand=3',
        'Formula': 'Safety Gap=-22, Available=0',
        'Example': 'Net GAP=-3, True GAP=0, Cause=Supply < Safety'
    })
    
    return pd.DataFrame(guide_data)


def _create_customer_sheet(customer_impact) -> pd.DataFrame:
    """Create customer impact sheet"""
    
    if customer_impact.customer_df.empty:
        return pd.DataFrame()
    
    columns_to_export = [
        'customer', 'customer_code', 'product_count',
        'total_required', 'total_shortage', 
        'total_demand_value', 'at_risk_value',
        'urgency'
    ]
    
    available_cols = [col for col in columns_to_export if col in customer_impact.customer_df.columns]
    customer_df = customer_impact.customer_df[available_cols].copy()
    
    column_rename = {
        'customer': 'Customer Name',
        'customer_code': 'Customer Code',
        'product_count': 'Products Affected',
        'total_required': 'Total Required Qty',
        'total_shortage': 'Total Shortage Qty',
        'total_demand_value': 'Demand Value (USD)',
        'at_risk_value': 'At Risk Value (USD)',
        'urgency': 'Urgency Level'
    }
    customer_df.rename(columns=column_rename, inplace=True)
    
    if 'At Risk Value (USD)' in customer_df.columns:
        customer_df = customer_df.sort_values('At Risk Value (USD)', ascending=False)
    
    return customer_df


def _format_excel_sheets(writer):
    """Apply formatting to Excel sheets"""
    
    try:
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
        
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Format header
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass
                
                adjusted_width = min(max(max_length + 2, 10), 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Freeze header
            worksheet.freeze_panes = 'A2'
            
    except ImportError:
        logger.warning("openpyxl styles not available, skipping formatting")


def export_gap_summary_csv(gap_df: pd.DataFrame) -> bytes:
    """Quick CSV export"""
    
    essential_cols = [
        'pt_code', 'product_name', 'brand',
        'total_supply', 'total_demand', 
        'safety_stock_qty', 'safety_gap', 'available_supply',
        'net_gap', 'true_gap', 'shortage_cause',
        'coverage_ratio', 'gap_status', 'priority'
    ]
    
    available_cols = [col for col in essential_cols if col in gap_df.columns]
    export_df = gap_df[available_cols].copy()
    
    output = io.BytesIO()
    export_df.to_csv(output, index=False, encoding='utf-8-sig')
    output.seek(0)
    
    return output.getvalue()